from __future__ import annotations

import io
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = ["Исходное название", "Стандартный заголовок", "Рекламный заголовок"]


def set_table_style(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    content_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    widths = {1: 68, 2: 58, 3: 88}
    for col_idx, width in widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = left
            cell.fill = content_fill
        ws.row_dimensions[row[0].row].height = 45


def read_source_excel(file_bytes: bytes, sheet_name: str = "Sheet1") -> list[str]:
    stream = io.BytesIO(file_bytes)
    wb = load_workbook(stream)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    titles: list[str] = []
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        titles.append(str(value) if value is not None else "")
    return titles


def generate_output_excel(
    origin_titles: Iterable[str],
    std_titles: Iterable[str],
    promo_titles: Iterable[str],
) -> bytes:
    origins = list(origin_titles)
    stds = list(std_titles)
    promos = list(promo_titles)
    if not (len(origins) == len(stds) == len(promos)):
        raise ValueError("origin_titles, std_titles, and promo_titles must have the same length")

    wb = Workbook()
    ws = wb.active
    ws.title = "Оптимизированные заголовки"

    for idx, (origin, standard, promo) in enumerate(zip(origins, stds, promos), start=2):
        ws.cell(row=idx, column=1, value=origin)
        ws.cell(row=idx, column=2, value=standard)
        ws.cell(row=idx, column=3, value=promo)

    set_table_style(ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_output_excel(
    output_path: str | Path,
    origin_titles: Iterable[str],
    std_titles: Iterable[str],
    promo_titles: Iterable[str],
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(generate_output_excel(origin_titles, std_titles, promo_titles))
    return path


if __name__ == "__main__":
    print("russian-title-optimizer Excel processor loaded")
    print("Use read_source_excel(), generate_output_excel(), or save_output_excel().")
